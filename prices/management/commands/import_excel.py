"""
Import crop price data from Excel files.

Supports two formats:
1. Ministry of Agriculture Crop Bulletin (mzigo.xlsx)
   - Columns: #, DATE, _, CROP, HIGH PRICE, LOW PRICE, UNIT, REGION, DISTRICT, MARKET
   - Prices per KG

2. Ministry of Industry & Trade Wholesale Price (.xls)
   - Prices in TZS per 100 KG
   - Region/District rows with Min/Max price columns per crop

Usage:
    python manage.py import_excel --mzigo "path/to/mzigo.xlsx"
    python manage.py import_excel --wholesale "path/to/Wholesale Price.xls"
    python manage.py import_excel --mzigo "path/to/mzigo.xlsx" --wholesale "path/to/Wholesale Price.xls"
"""
import os
import re
from datetime import datetime, date
from django.core.management.base import BaseCommand
from prices.models import Region, Market, Crop, PriceEntry

SWAHILI_TO_ENGLISH = {
    'Mahindi': ('grain', 'Maize', 'kg'),
    'Mchele': ('grain', 'Rice', 'kg'),
    'Mpunga': ('grain', 'Rice', 'kg'),
    'Maharage': ('legume', 'Beans', 'kg'),
    'Mtama': ('grain', 'Sorghum', 'kg'),
    'Ulezi': ('grain', 'Finger Millet', 'kg'),
    'Uwele': ('grain', 'Bulrush Millet', 'kg'),
    'Ngano': ('grain', 'Wheat', 'kg'),
    'Alizeti': ('cash', 'Sunflower', 'kg'),
    'Karanga': ('legume', 'Groundnuts', 'kg'),
    'Njugu Mawe': ('legume', 'Bambara Groundnuts', 'kg'),
    'Mbaazi': ('legume', 'Pigeon Peas', 'kg'),
    'Choroko': ('legume', 'Green Gram', 'kg'),
    'Kunde': ('legume', 'Cowpeas', 'kg'),
    'Viazi Mviringo': ('root', 'Irish Potatoes', 'kg'),
    'Viazi Vitamu': ('root', 'Sweet Potatoes', 'kg'),
    'Viazi': ('root', 'Irish Potatoes', 'kg'),
    'Mihogo': ('root', 'Cassava', 'kg'),
    'Ndizi': ('fruit', 'Bananas', 'kg'),
    'Nyanya': ('vegetable', 'Tomatoes', 'kg'),
    'Vitunguu': ('vegetable', 'Onions', 'kg'),
    'Vitunguu Swaumu': ('vegetable', 'Garlic', 'kg'),
    'Hoho': ('vegetable', 'Bell Peppers', 'kg'),
    'Kabichi': ('vegetable', 'Cabbage', 'kg'),
    'Karoti': ('vegetable', 'Carrots', 'kg'),
    'Bamia': ('vegetable', 'Okra', 'kg'),
    'Maboga': ('vegetable', 'Pumpkin', 'kg'),
    'Tango': ('vegetable', 'Cucumber', 'kg'),
    'Njegere': ('legume', 'Peas', 'kg'),
    'Dengu': ('legume', 'Lentils', 'kg'),
    'Figiri': ('legume', 'Pigeon Peas', 'kg'),
    'Maembe': ('fruit', 'Mangoes', 'kg'),
    'Nanasi': ('fruit', 'Pineapples', 'piece'),
    'Tikiti Maji': ('fruit', 'Watermelon', 'piece'),
    'Parachichi': ('fruit', 'Avocado', 'kg'),
    'Mapapai': ('fruit', 'Papaya', 'kg'),
    'Chenza': ('fruit', 'Tangerine', 'kg'),
    'Pasheni': ('fruit', 'Passion Fruit', 'kg'),
    'Chainizi': ('fruit', 'Oranges', 'kg'),
    'Magimbi': ('root', 'Taro', 'kg'),
    'Tangawizi': ('spice', 'Ginger', 'kg'),
    'Bilinganya': ('vegetable', 'Eggplant', 'kg'),
    'Mchicha': ('vegetable', 'Amaranth', 'bundle'),
    'Pilipili': ('spice', 'Chili Pepper', 'kg'),
    'Soya': ('legume', 'Soybeans', 'kg'),
    'Ufuta': ('cash', 'Sesame', 'kg'),
    'Pamba': ('cash', 'Cotton', 'kg'),
    'Kahawa': ('cash', 'Coffee', 'kg'),
    'Korosho': ('cash', 'Cashew Nuts', 'kg'),
    'Karafuu': ('spice', 'Cloves', 'kg'),
    'Tumbaku': ('cash', 'Tobacco', 'kg'),
    'Chai': ('cash', 'Tea', 'kg'),
    'Limau': ('fruit', 'Lemons', 'kg'),
    'Mapera': ('fruit', 'Guavas', 'kg'),
    'Machungwa': ('fruit', 'Oranges', 'kg'),
}

REGION_ZONES = {
    'ARUSHA': 'Northern',
    'KILIMANJARO': 'Northern',
    'TANGA': 'Northern',
    'MANYARA': 'Northern',
    'DAR ES SALAAM': 'Coastal',
    'PWANI': 'Coastal',
    'MOROGORO': 'Coastal',
    'DODOMA': 'Central',
    'SINGIDA': 'Central',
    'TABORA': 'Central',
    'MBEYA': 'Southern Highlands',
    'IRINGA': 'Southern Highlands',
    'NJOMBE': 'Southern Highlands',
    'SONGWE': 'Southern Highlands',
    'RUVUMA': 'Southern',
    'LINDI': 'Southern',
    'MTWARA': 'Southern',
    'MARA': 'Lake',
    'MWANZA': 'Lake',
    'SHINYANGA': 'Lake',
    'KAGERA': 'Lake',
    'GEITA': 'Lake',
    'SIMIYU': 'Lake',
    'KIGOMA': 'Western',
    'KATAVI': 'Western',
    'RUKWA': 'Western',
}


REGION_SPELLING_CORRECTIONS = {
    'DAR ES SAALAM': 'DAR ES SALAAM',
    'DAR ES SALAAM': 'DAR ES SALAAM',
    'DARESSALAAM': 'DAR ES SALAAM',
    'MBEYA ': 'MBEYA',
    'SINGIDA ': 'SINGIDA',
    'DODOMA ': 'DODOMA',
}


def normalize_region(name):
    if not name:
        return None
    name = name.strip().upper()
    corrected = REGION_SPELLING_CORRECTIONS.get(name, name)
    if corrected in REGION_ZONES:
        return corrected.title()
    return None


def get_or_create_region(name):
    if not name:
        return None
    nname = normalize_region(name)
    if not nname:
        return None
    zone = REGION_ZONES.get(nname.upper(), 'General')
    obj, _ = Region.objects.get_or_create(name=nname, defaults={'zone': zone})
    return obj


def get_or_create_market(name, region, district=None):
    if not name or not region:
        return None
    mname = name.strip().title()
    obj, _ = Market.objects.get_or_create(
        name=mname, region=region,
        defaults={'district': district.strip().title() if district else ''}
    )
    return obj


def map_crop(swahili_name):
    if not swahili_name:
        return None
    sname = swahili_name.strip().title()
    mapping = SWAHILI_TO_ENGLISH.get(sname)
    if mapping:
        return mapping
    for key, val in SWAHILI_TO_ENGLISH.items():
        if key.lower() == sname.lower():
            return val
    return None


def get_or_create_crop(swahili_name):
    mapping = map_crop(swahili_name)
    if not mapping:
        return None
    cat, eng_name, unit = mapping
    obj, _ = Crop.objects.get_or_create(
        name=eng_name,
        defaults={'category': cat, 'unit': unit}
    )
    return obj


def parse_date_dmy(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    try:
        parts = date_str.split('-')
        if len(parts) == 3:
            return date(int(parts[2]), int(parts[1]), int(parts[0]))
    except (ValueError, IndexError):
        pass
    try:
        if isinstance(date_str, (int, float)):
            from datetime import timedelta
            base = date(1900, 1, 1)
            return base + timedelta(days=int(date_str) - 2)
    except (ValueError, TypeError):
        pass
    return None


class Command(BaseCommand):
    help = 'Import crop price data from Excel files (mzigo.xlsx and Wholesale Price .xls)'

    def add_arguments(self, parser):
        parser.add_argument('--mzigo', type=str, help='Path to mzigo.xlsx file (Crop Bulletin format)')
        parser.add_argument('--wholesale', type=str, help='Path to Wholesale Price .xls file')

    def handle(self, *args, **options):
        mzigo_path = options.get('mzigo')
        wholesale_path = options.get('wholesale')

        if not mzigo_path and not wholesale_path:
            self.stderr.write(self.style.ERROR('Provide at least --mzigo or --wholesale'))
            return

        if mzigo_path:
            if not os.path.exists(mzigo_path):
                self.stderr.write(self.style.ERROR(f'mzigo file not found: {mzigo_path}'))
            else:
                self._import_mzigo(mzigo_path)

        if wholesale_path:
            if not os.path.exists(wholesale_path):
                self.stderr.write(self.style.ERROR(f'Wholesale file not found: {wholesale_path}'))
            else:
                self._import_wholesale(wholesale_path)

        total = PriceEntry.objects.count()
        self.stdout.write(self.style.SUCCESS(f'\nTotal price entries now: {total}'))

    def _import_mzigo(self, path):
        self.stdout.write(self.style.SUCCESS(f'\nImporting mzigo.xlsx from: {path}'))
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        stats = {'total': 0, 'created': 0, 'skipped': 0, 'new_crops': 0, 'new_regions': 0, 'new_markets': 0}
        errors = []

        for row_idx in range(2, ws.max_row + 1):
            row_num = row_idx + 1
            try:
                date_val = ws.cell(row_idx, 2).value
                crop_name = ws.cell(row_idx, 4).value
                high_price = ws.cell(row_idx, 5).value
                low_price = ws.cell(row_idx, 6).value
                unit = ws.cell(row_idx, 7).value
                region_name = ws.cell(row_idx, 8).value
                district_name = ws.cell(row_idx, 9).value
                market_name = ws.cell(row_idx, 10).value

                if not crop_name or not region_name:
                    continue

                crop = get_or_create_crop(crop_name)
                if not crop:
                    errors.append(f'Row {row_num}: Unknown crop "{crop_name}"')
                    stats['skipped'] += 1
                    continue

                region = get_or_create_region(region_name)
                if not region:
                    errors.append(f'Row {row_num}: Unknown region "{region_name}"')
                    stats['skipped'] += 1
                    continue

                market = get_or_create_market(market_name, region, district_name)
                if not market:
                    market = Market.objects.filter(region=region).first()
                    if not market:
                        errors.append(f'Row {row_num}: No market for region {region.name}')
                        stats['skipped'] += 1
                        continue

                price_date = parse_date_dmy(date_val)
                if not price_date:
                    errors.append(f'Row {row_num}: Invalid date "{date_val}"')
                    stats['skipped'] += 1
                    continue

                if isinstance(high_price, (int, float)) and high_price > 0:
                    PriceEntry.objects.get_or_create(
                        crop=crop, market=market, price_date=price_date, price=float(high_price),
                        defaults={'status': 'approved'}
                    )
                    stats['created'] += 1

                if isinstance(low_price, (int, float)) and low_price > 0:
                    PriceEntry.objects.get_or_create(
                        crop=crop, market=market, price_date=price_date, price=float(low_price),
                        defaults={'status': 'approved'}
                    )
                    stats['created'] += 1

                stats['total'] += 1

            except Exception as e:
                errors.append(f'Row {row_num}: {e}')
                stats['skipped'] += 1

        self.stdout.write(self.style.SUCCESS(
            f'mzigo: {stats["created"]} prices created from {stats["total"]} rows, '
            f'{stats["skipped"]} skipped'
        ))
        if errors:
            self.stdout.write(self.style.WARNING(f'Errors ({len(errors)}):'))
            for e in errors[:20]:
                self.stdout.write(f'  {e}')
            if len(errors) > 20:
                self.stdout.write(f'  ... and {len(errors) - 20} more')

    def _import_wholesale(self, path):
        self.stdout.write(self.style.SUCCESS(f'\nImporting Wholesale Price from: {path}'))
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_name('WP001')

        crop_cols = []
        header_row = 1
        for c in range(2, ws.ncols, 2):
            crop_label = str(ws.cell_value(header_row, c)).strip()
            if crop_label and crop_label != '   ':
                match = re.match(r'^([A-Za-z\s]+)', crop_label)
                if match:
                    crop_cols.append((c, match.group(1).strip()))
                else:
                    crop_cols.append((c, crop_label))
            else:
                break

        stats = {'created': 0, 'skipped': 0}
        errors = []

        for row_idx in range(3, ws.nrows):
            try:
                region_name = str(ws.cell_value(row_idx, 0)).strip()
                market_name = str(ws.cell_value(row_idx, 1)).strip()

                if not region_name or region_name.startswith('CHANZO') or region_name.startswith('IDARA') or region_name == '':
                    continue

                region = get_or_create_region(region_name)
                if not region:
                    errors.append(f'Row {row_idx}: Unknown region "{region_name}"')
                    stats['skipped'] += 1
                    continue

                market = get_or_create_market(market_name, region)
                if not market:
                    errors.append(f'Row {row_idx}: No market for {region_name}/{market_name}')
                    stats['skipped'] += 1
                    continue

                for col_offset, crop_label in crop_cols:
                    min_val = ws.cell_value(row_idx, col_offset)
                    max_val = ws.cell_value(row_idx, col_offset + 1)

                    eng_name = crop_label.split('(')[0].strip()

                    crop = Crop.objects.filter(name__iexact=eng_name).first()
                    if not crop:
                        mapping = map_crop(eng_name)
                        if mapping:
                            crop = get_or_create_crop(eng_name)
                    if not crop:
                        continue

                    price_date = date(2026, 6, 1)

                    if isinstance(min_val, (int, float)) and min_val > 0:
                        price_per_kg = min_val / 100.0
                        PriceEntry.objects.get_or_create(
                            crop=crop, market=market, price_date=price_date, price=round(price_per_kg, 2),
                            defaults={'status': 'approved'}
                        )
                        stats['created'] += 1

                    if isinstance(max_val, (int, float)) and max_val > 0:
                        price_per_kg = max_val / 100.0
                        PriceEntry.objects.get_or_create(
                            crop=crop, market=market, price_date=price_date, price=round(price_per_kg, 2),
                            defaults={'status': 'approved'}
                        )
                        stats['created'] += 1

            except Exception as e:
                errors.append(f'Row {row_idx}: {e}')
                stats['skipped'] += 1

        self.stdout.write(self.style.SUCCESS(
            f'Wholesale: {stats["created"]} prices created, {stats["skipped"]} skipped'
        ))
        if errors:
            self.stdout.write(self.style.WARNING(f'Errors ({len(errors)}):'))
            for e in errors[:20]:
                self.stdout.write(f'  {e}')
