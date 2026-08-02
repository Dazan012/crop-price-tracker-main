import csv
import io
from datetime import date, datetime

from django.http import HttpResponse
from django.utils import timezone

from .models import PriceEntry, Crop, Market, Region


def get_price_queryset(params):
    """Build filtered PriceEntry queryset from request params."""
    qs = PriceEntry.objects.select_related('crop', 'market__region', 'submitted_by').filter(status='approved')

    crop_id = params.get('crop')
    market_id = params.get('market')
    region_id = params.get('region')
    date_from = params.get('date_from')
    date_to = params.get('date_to')

    if crop_id:
        qs = qs.filter(crop_id=crop_id)
    if market_id:
        qs = qs.filter(market_id=market_id)
    if region_id:
        qs = qs.filter(market__region_id=region_id)
    if date_from:
        qs = qs.filter(price_date__gte=date_from)
    if date_to:
        qs = qs.filter(price_date__lte=date_to)

    return qs.order_by('-price_date', 'crop__name', 'market__name')


def get_report_meta(params):
    """Build metadata dict for report headers."""
    meta = {
        'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'report_type': 'Price Report',
    }
    if params.get('crop'):
        try:
            meta['crop'] = Crop.objects.get(id=params['crop']).name
        except Crop.DoesNotExist:
            pass
    if params.get('market'):
        try:
            meta['market'] = Market.objects.get(id=params['market']).name
        except Market.DoesNotExist:
            pass
    if params.get('region'):
        try:
            meta['region'] = Region.objects.get(id=params['region']).name
        except Region.DoesNotExist:
            pass
    if params.get('date_from'):
        meta['date_from'] = params['date_from']
    if params.get('date_to'):
        meta['date_to'] = params['date_to']
    return meta


def row_data(entry):
    """Convert a PriceEntry to a tuple for export."""
    return (
        entry.crop.name,
        entry.market.name,
        entry.market.region.name,
        entry.price,
        entry.quantity or '',
        entry.price_date.isoformat(),
        entry.submitted_by.username if entry.submitted_by else 'N/A',
        entry.submitted_at.strftime('%Y-%m-%d %H:%M') if entry.submitted_at else '',
    )


HEADERS = ['Crop', 'Market', 'Region', 'Price (TZS)', 'Quantity', 'Date', 'Submitted By', 'Submitted At']


def generate_csv(params):
    """Generate CSV report. Returns HttpResponse with attachment."""
    qs = get_price_queryset(params)
    meta = get_report_meta(params)

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Metadata header
    for k, v in meta.items():
        writer.writerow([f'# {k.replace("_", " ").title()}: {v}'])
    writer.writerow([])
    writer.writerow(HEADERS)

    for entry in qs.iterator(chunk_size=500):
        writer.writerow(row_data(entry))

    csv_content = buffer.getvalue()
    buffer.close()

    filename = f"price_report_{date.today().isoformat()}.csv"
    response = HttpResponse(csv_content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_excel(params):
    """Generate Excel (.xlsx) report. Requires openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    qs = get_price_queryset(params)
    meta = get_report_meta(params)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Price Report'

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    meta_font = Font(name='Calibri', italic=True, color='555555', size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Metadata rows
    row_idx = 1
    for k, v in meta.items():
        ws.cell(row=row_idx, column=1, value=f'{k.replace("_", " ").title()}: {v}').font = meta_font
        row_idx += 1

    row_idx += 1  # blank row

    # Header row
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    row_idx += 1

    # Data rows
    for entry in qs.iterator(chunk_size=500):
        data = row_data(entry)
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx == 4:  # Price column
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        row_idx += 1

    # Auto-fit columns
    for col_idx, header in enumerate(HEADERS, 1):
        max_len = len(header)
        for r in range(row_idx - len(qs), row_idx):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"price_report_{date.today().isoformat()}.xlsx"
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_pdf(params):
    """Generate PDF report. Requires reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                     Table, TableStyle, PageBreak)

    qs = get_price_queryset(params)
    meta = get_report_meta(params)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=6)
    meta_style = ParagraphStyle('MetaInfo', parent=styles['Normal'], fontSize=8, textColor=colors.grey, spaceAfter=2)

    elements = []

    # Title
    elements.append(Paragraph('Crop Price Report', title_style))
    elements.append(Spacer(1, 4 * mm))

    # Metadata
    for k, v in meta.items():
        elements.append(Paragraph(f'<b>{k.replace("_", " ").title()}:</b> {v}', meta_style))
    elements.append(Spacer(1, 6 * mm))

    # Table
    table_data = [HEADERS]
    for entry in qs.iterator(chunk_size=500):
        table_data.append(list(row_data(entry)))

    col_widths = [50*mm, 45*mm, 35*mm, 25*mm, 20*mm, 25*mm, 30*mm, 30*mm]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.106, 0.369, 0.125)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.97, 0.93)]),
    ]))
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        f'Generated by Smart Crops Market Tracker on {timezone.now().strftime("%Y-%m-%d %H:%M")}',
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=1)
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"price_report_{date.today().isoformat()}.pdf"
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_summary_csv(params):
    """Generate a summary/aggregate CSV report."""
    qs = get_price_queryset(params)

    from django.db.models import Avg, Min, Max, Count
    summary = qs.values('crop__name', 'market__name', 'market__region__name').annotate(
        avg_price=Avg('price'),
        min_price=Min('price'),
        max_price=Max('price'),
        entry_count=Count('id'),
    ).order_by('crop__name', 'market__region__name')

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow(['# Report: Price Summary'])
    writer.writerow([f'# Generated: {timezone.now().strftime("%Y-%m-%d %H:%M")}'])
    writer.writerow([])
    writer.writerow(['Crop', 'Market', 'Region', 'Avg Price', 'Min Price', 'Max Price', 'Entries'])

    for row in summary:
        writer.writerow([
            row['crop__name'],
            row['market__name'],
            row['market__region__name'],
            round(row['avg_price'], 0) if row['avg_price'] else '',
            row['min_price'] or '',
            row['max_price'] or '',
            row['entry_count'],
        ])

    csv_content = buffer.getvalue()
    buffer.close()

    filename = f"price_summary_{date.today().isoformat()}.csv"
    response = HttpResponse(csv_content, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
